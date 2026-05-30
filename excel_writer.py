from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

C_COLORS = {
    'J&T Express':'D6EAFF','GTL':'D6F5E3','SiCepat':'FFF0CC',
    'Shopee Express':'FCE4D6','Wahana':'E8D5F5','Anteraja':'D5ECF5',
    'LEX':'FFD6D6','Unknown':'F2F2F2',
}
HDR  = PatternFill('solid',fgColor='1B1F2E')
HFNT = Font(bold=True,color='FFFFFF',size=10,name='Calibri')
C    = Alignment(horizontal='center',vertical='center',wrap_text=True)
L    = Alignment(horizontal='left',vertical='center',wrap_text=True)
BD   = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
              top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))

def H(ws,cols):  # header row
    for i,(lbl,w) in enumerate(cols,1):
        c=ws.cell(row=1,column=i,value=lbl)
        c.fill=HDR;c.font=HFNT;c.alignment=C;c.border=BD
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[1].height=30;ws.freeze_panes='A2'

def R(ws,row,vals,fgcolor,aligns):
    f=PatternFill('solid',fgColor=fgcolor)
    for i,(v,al) in enumerate(zip(vals,aligns),1):
        c=ws.cell(row=row,column=i,value=v)
        c.fill=f;c.alignment=al;c.border=BD;c.font=Font(size=9,name='Calibri')

def write_excel(rows,output_path):
    wb=Workbook()

    # ── SHEET 1: Semua Resi ───────────────────────────────────────────────────
    ws1=wb.active;ws1.title="Semua Resi"
    H(ws1,[('No',5),('Kurir',15),('Layanan',9),('No. Resi',22),
           ('Nama Produk',34),('Seller SKU',24),('Variasi',16),
           ('Qty',6),('Nama Penerima',22),('Alamat',42),('Platform',14)])

    srows=sorted(rows,key=lambda r:(r['courier'],r['no_resi']))
    prev=None;flip=True
    for i,r in enumerate(srows,1):
        if r['no_resi']!=prev: flip=not flip; prev=r['no_resi']
        base=C_COLORS.get(r['courier'],'F2F2F2')
        clr=base if flip else base[:1]+'F'+base[2:]
        nama=(r['nama_produk'] or '').rstrip(',').strip()
        R(ws1,i+1,[i,r['courier'],r['layanan'],r['no_resi'],
                   nama,r['sku'],r['variasi'] or '-',r['qty'],
                   r['nama_penerima'],r['alamat'],r['platform']],
          clr,[C,L,C,L,L,L,C,C,L,L,L])
        if r['qty']>1:
            ws1.cell(row=i+1,column=8).font=Font(bold=True,color='CC0000',size=11,name='Calibri')
    ws1.auto_filter.ref=f"A1:{get_column_letter(11)}1"

    # ── SHEET 2: Rekap Kurir — RINGKAS & JELAS ────────────────────────────────
    # Satu baris per kurir, total resi + total qty, mudah dibaca
    ws2=wb.create_sheet("Rekap Kurir")
    H(ws2,[('Kurir',18),('Total Resi',14),('Total Qty',12),
           ('Platform',18),('Layanan',12)])

    by_courier=defaultdict(lambda:{'resi':set(),'qty':0,'platform':set(),'layanan':set()})
    for r in rows:
        by_courier[r['courier']]['resi'].add(r['no_resi'])
        by_courier[r['courier']]['qty']+=r['qty']
        by_courier[r['courier']]['platform'].add(r['platform'])
        by_courier[r['courier']]['layanan'].add(r['layanan'])

    for i,(cur,v) in enumerate(sorted(by_courier.items()),2):
        clr=C_COLORS.get(cur,'F2F2F2')
        R(ws2,i,[cur,len(v['resi']),v['qty'],
                 ', '.join(sorted(v['platform'])),
                 ', '.join(sorted(v['layanan']))],
          clr,[L,C,C,L,L])
        ws2.cell(row=i,column=2).font=Font(bold=True,size=11,name='Calibri')
        ws2.cell(row=i,column=3).font=Font(bold=True,size=11,name='Calibri')

    tr=len(by_courier)+2
    tf=PatternFill('solid',fgColor='F5A623')
    for col,val in enumerate(['TOTAL',
        len(set(r['no_resi'] for r in rows)),
        sum(r['qty'] for r in rows),'',''],1):
        c=ws2.cell(row=tr,column=col,value=val)
        c.fill=tf;c.font=Font(bold=True,size=11,name='Calibri')
        c.alignment=C;c.border=BD

    # ── SHEET 3: Rekap SKU — stok yang harus disiapkan ────────────────────────
    ws3=wb.create_sheet("Rekap SKU")
    H(ws3,[('Seller SKU',26),('Variasi / Ukuran',18),
           ('Nama Produk',32),('Total Resi',12),('Total Qty',12)])

    sku_s=defaultdict(lambda:{'resi':set(),'qty':0,'nama':'','variasi':''})
    for r in rows:
        k=r['sku'] or '(tanpa SKU)'
        sku_s[k]['resi'].add(r['no_resi'])
        sku_s[k]['qty']+=r['qty']
        if not sku_s[k]['nama'] and r.get('nama_produk'):
            sku_s[k]['nama']=r['nama_produk'].rstrip(',').strip()
        if not sku_s[k]['variasi'] and r.get('variasi'):
            sku_s[k]['variasi']=r['variasi']

    gf=PatternFill('solid',fgColor='EAF3DE')
    for i,(sku,v) in enumerate(sorted(sku_s.items(),key=lambda x:-x[1]['qty']),2):
        R(ws3,i,[sku,v['variasi'],v['nama'],len(v['resi']),v['qty']],
          'EAF3DE',[L,C,L,C,C])
        ws3.cell(row=i,column=5).font=Font(bold=True,size=11,name='Calibri')

    # Total
    tr3=len(sku_s)+2
    for col,val in enumerate(['TOTAL','','',
        len(set(r['no_resi'] for r in rows)),
        sum(r['qty'] for r in rows)],1):
        c=ws3.cell(row=tr3,column=col,value=val)
        c.fill=tf;c.font=Font(bold=True,size=11,name='Calibri')
        c.alignment=C;c.border=BD

    # ── SHEET 4: Packing List Gudang ─────────────────────────────────────────
    # Diurutkan per resi. Multi produk 1 resi = beberapa baris warna sama.
    # Qty > 1 = bold merah. Kolom ceklis kosong untuk tanda tangan gudang.
    ws4=wb.create_sheet("Packing List Gudang")
    H(ws4,[('No Urut',7),('No. Resi',22),('Kurir',13),
           ('Nama Produk',34),('Seller SKU',24),('Variasi',14),
           ('Qty',6),('✓ Ambil',8)])

    # Kelompokkan per resi
    resi_groups=defaultdict(list)
    for r in srows: resi_groups[r['no_resi']].append(r)

    row_num=2; no_urut=1
    MULTI_COLORS=['FFF9C4','E8F5E9','E3F2FD','FCE4D6','F3E5F5','E0F7FA']
    single_colors=iter(['F5F5F5','FFFFFF']*300)
    mc_idx=0

    for resi,items in resi_groups.items():
        if len(items)>1:
            # Multi produk: semua baris warna yang sama (kuning/hijau dll bergantian)
            grp_color=MULTI_COLORS[mc_idx % len(MULTI_COLORS)]; mc_idx+=1
            for r in items:
                nama=(r['nama_produk'] or '').rstrip(',').strip()
                R(ws4,row_num,[no_urut,r['no_resi'],r['courier'],
                               nama,r['sku'],r['variasi'] or '-',r['qty'],''],
                  grp_color,[C,L,L,L,L,C,C,C])
                if r['qty']>1:
                    ws4.cell(row=row_num,column=7).font=Font(bold=True,color='CC0000',size=11,name='Calibri')
                row_num+=1; no_urut+=1
            # Garis pemisah setelah grup multi
            for col in range(1,9):
                ws4.cell(row=row_num-1,column=col).border=Border(
                    left=Side(style='thin',color='CCCCCC'),
                    right=Side(style='thin',color='CCCCCC'),
                    top=Side(style='thin',color='CCCCCC'),
                    bottom=Side(style='medium',color='888888'))
        else:
            r=items[0]
            clr=next(single_colors)
            nama=(r['nama_produk'] or '').rstrip(',').strip()
            R(ws4,row_num,[no_urut,r['no_resi'],r['courier'],
                           nama,r['sku'],r['variasi'] or '-',r['qty'],''],
              clr,[C,L,L,L,L,C,C,C])
            if r['qty']>1:
                ws4.cell(row=row_num,column=7).font=Font(bold=True,color='CC0000',size=11,name='Calibri')
            row_num+=1; no_urut+=1

    # ── SHEET 5: Ringkasan Gudang (tampilan paling simpel) ────────────────────
    # Satu halaman: hari ini siapkan berapa per SKU, berapa resi multi-produk
    ws5=wb.create_sheet("Ringkasan Hari Ini")
    ws5.column_dimensions['A'].width=30
    ws5.column_dimensions['B'].width=18
    ws5.column_dimensions['C'].width=14

    # Judul
    ws5.merge_cells('A1:C1')
    jdl=ws5.cell(row=1,column=1,value='📦  RINGKASAN PRODUK HARI INI')
    jdl.font=Font(bold=True,size=14,name='Calibri',color='1B1F2E')
    jdl.alignment=C
    jdl.fill=PatternFill('solid',fgColor='F5A623')

    # Header tabel stok
    ws5.merge_cells('A2:C2')
    h2=ws5.cell(row=2,column=1,value='STOK YANG HARUS DISIAPKAN')
    h2.font=Font(bold=True,size=10,name='Calibri',color='FFFFFF')
    h2.fill=HDR;h2.alignment=C

    for col,lbl in enumerate(['Produk / Seller SKU','Variasi','Qty Disiapkan'],1):
        c=ws5.cell(row=3,column=col,value=lbl)
        c.fill=HDR;c.font=HFNT;c.alignment=C;c.border=BD

    gf2=PatternFill('solid',fgColor='EAF3DE')
    row5=4
    for sku,v in sorted(sku_s.items(),key=lambda x:-x[1]['qty']):
        for col,val in enumerate([sku,v['variasi'],v['qty']],1):
            c=ws5.cell(row=row5,column=col,value=val)
            c.fill=gf2;c.alignment=L if col<3 else C;c.border=BD
            c.font=Font(size=10,name='Calibri',bold=(col==3))
        row5+=1
    # Total stok
    for col,val in enumerate(['TOTAL PRODUK','',sum(r['qty'] for r in rows)],1):
        c=ws5.cell(row=row5,column=col,value=val)
        c.fill=tf;c.font=Font(bold=True,size=11,name='Calibri');c.alignment=C;c.border=BD
    row5+=2

    # Resi multi produk
    ws5.merge_cells(f'A{row5}:C{row5}')
    h3=ws5.cell(row=row5,column=1,value='RESI DENGAN LEBIH DARI 1 PRODUK / VARIASI')
    h3.font=Font(bold=True,size=10,name='Calibri',color='FFFFFF')
    h3.fill=PatternFill('solid',fgColor='CC0000');h3.alignment=C
    row5+=1

    for col,lbl in enumerate(['No. Resi','Produk & Qty','Kurir'],1):
        c=ws5.cell(row=row5,column=col,value=lbl)
        c.fill=HDR;c.font=HFNT;c.alignment=C;c.border=BD
    row5+=1

    mf=PatternFill('solid',fgColor='FFF9C4')
    multi_resi={resi:items for resi,items in resi_groups.items() if len(items)>1}
    for resi,items in sorted(multi_resi.items()):
        detail=' | '.join(f"{(r['sku'] or r['nama_produk'])[:15]} x{r['qty']}" for r in items)
        for col,val in enumerate([resi,detail,items[0]['courier']],1):
            c=ws5.cell(row=row5,column=col,value=val)
            c.fill=mf;c.alignment=L;c.border=BD
            c.font=Font(size=9,name='Calibri')
        row5+=1

    ws5.column_dimensions['B'].width=45

    wb.save(output_path)
    return output_path

def write_excel_multi(rows, output_path):
    """
    Versi multi-akun: rows sudah punya field 'akun' dan 'waktu'
    Tambah Sheet: Rekap Gudang (per akun per waktu per produk)
    """
    # Panggil write_excel biasa dulu
    from openpyxl import load_workbook
    write_excel(rows, output_path)
    wb = load_workbook(output_path)

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    HDR2  = PatternFill('solid',fgColor='1B1F2E')
    HFNT2 = Font(bold=True,color='FFFFFF',size=10,name='Calibri')
    C2    = Alignment(horizontal='center',vertical='center',wrap_text=True)
    L2    = Alignment(horizontal='left',vertical='center',wrap_text=True)
    BD2   = Border(left=Side(style='thin',color='CCCCCC'),
                   right=Side(style='thin',color='CCCCCC'),
                   top=Side(style='thin',color='CCCCCC'),
                   bottom=Side(style='thin',color='CCCCCC'))
    AMBER = PatternFill('solid',fgColor='F5A623')
    RED_F = PatternFill('solid',fgColor='FFEBEE')
    RED_T = Font(color='CC0000',size=8,italic=True,name='Calibri')

    HSD_F = PatternFill('solid',fgColor='1A3A5C')
    HSS_F = PatternFill('solid',fgColor='5C2A00')
    HSD_ROW = {'PAGI':'D6E4FF','SIANG':'C5D8F5','SORE':'B8CAE8'}
    HSS_ROW = {'PAGI':'FFE4CC','SIANG':'F5D4B8','SORE':'E8C4A8'}

    WAKTU_ORDER = ['PAGI','SIANG','SORE']

    # ── SHEET: Rekap Gudang ───────────────────────────────────────────────────
    ws = wb.create_sheet("Rekap Gudang", 0)  # taruh di posisi pertama
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10

    # Judul
    ws.merge_cells('A1:E1')
    jdl = ws.cell(row=1,column=1,value=f'📦  REKAP GUDANG — {rows[0].get("akun","") if rows else ""}')
    jdl.font = Font(bold=True,size=13,name='Calibri',color='FFFFFF')
    jdl.fill = PatternFill('solid',fgColor='1B1F2E')
    jdl.alignment = C2

    # Header
    for col,(lbl,_) in enumerate([('Akun',10),('Waktu',12),
                                    ('Nama Produk / SKU',32),
                                    ('Variasi',18),('Total Qty',10)],1):
        c = ws.cell(row=2,column=col,value=lbl)
        c.fill=HDR2; c.font=HFNT2; c.alignment=C2; c.border=BD2
    ws.row_dimensions[2].height=26
    ws.freeze_panes='A3'

    # Hitung per akun x waktu x sku
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    nama_map = {}; variasi_map = {}
    for r in rows:
        akun = r.get('akun','HSD')
        waktu = r.get('waktu','PAGI')
        sku = r['sku'] or '(cek manual)'
        data[akun][waktu][sku] += r['qty']
        if sku not in nama_map and r.get('nama_produk'):
            nama_map[sku] = r['nama_produk'].rstrip(',').strip()
        if sku not in variasi_map and r.get('variasi'):
            variasi_map[sku] = r['variasi']

    row_n = 3
    for akun in ['HSD','HSS']:
        if akun not in data: continue
        akun_color = HSD_F if akun=='HSD' else HSS_F
        row_colors = HSD_ROW if akun=='HSD' else HSS_ROW

        for waktu in WAKTU_ORDER:
            if waktu not in data[akun]: continue
            sku_qty = sorted(data[akun][waktu].items(), key=lambda x: -x[1])

            for sku, qty in sku_qty:
                nama = nama_map.get(sku,'')
                variasi = variasi_map.get(sku,'')
                clr = row_colors[waktu]
                f = PatternFill('solid',fgColor=clr)

                vals = [akun, waktu, nama or sku, variasi, qty]
                aligns = [C2,C2,L2,C2,C2]
                for col,(v,al) in enumerate(zip(vals,aligns),1):
                    c = ws.cell(row=row_n,column=col,value=v)
                    c.fill=f; c.alignment=al; c.border=BD2
                    c.font=Font(size=10,name='Calibri')

                # Akun kolom: warna khusus bold
                c_akun = ws.cell(row=row_n,column=1)
                c_akun.fill = akun_color
                c_akun.font = Font(bold=True,color='FFFFFF',size=10,name='Calibri')

                # Variasi kosong = memang tidak ada di PDF (GTL pakai Default)
                if not variasi:
                    variasi = '-'

                # Kalau kosong beri keterangan merah
                if not nama:
                    c_nama = ws.cell(row=row_n,column=3)
                    c_nama.value = f'{sku}  ← nama tidak terbaca di PDF'
                    c_nama.font = Font(color='CC0000',size=9,italic=True,name='Calibri')

                row_n += 1

            # Subtotal per akun-waktu
            subtotal = sum(q for _,q in sku_qty)
            ws.merge_cells(f'A{row_n}:D{row_n}')
            c = ws.cell(row=row_n,column=1,
                        value=f'  SUBTOTAL {akun} {waktu}')
            c.fill=AMBER; c.font=Font(bold=True,size=10,name='Calibri')
            c.alignment=L2; c.border=BD2
            ct = ws.cell(row=row_n,column=5,value=subtotal)
            ct.fill=AMBER; ct.font=Font(bold=True,size=12,name='Calibri')
            ct.alignment=C2; ct.border=BD2
            row_n += 1

        # Spasi antar akun
        row_n += 1

    # GRAND TOTAL
    ws.merge_cells(f'A{row_n}:D{row_n}')
    c=ws.cell(row=row_n,column=1,value='  GRAND TOTAL SEMUA')
    c.fill=PatternFill('solid',fgColor='1B1F2E')
    c.font=Font(bold=True,color='FFFFFF',size=11,name='Calibri'); c.alignment=L2; c.border=BD2
    ct=ws.cell(row=row_n,column=5,value=sum(r['qty'] for r in rows))
    ct.fill=PatternFill('solid',fgColor='1B1F2E')
    ct.font=Font(bold=True,color='FFFFFF',size=13,name='Calibri'); ct.alignment=C2; ct.border=BD2

    # ── SHEET: Ringkasan Order ────────────────────────────────────────────────
    # Per Akun+Waktu: kelompokkan resi berdasarkan KOMBINASI produk yang dibeli
    # Contoh: "BG 220gr x1" → 150 resi | "BG 220gr x2" → 30 resi | "BG 220gr + Drink" → 10 resi
    from collections import Counter, defaultdict as dd2
    ws6 = wb.create_sheet("Ringkasan Order")
    ws6.freeze_panes='A3'
    for w,width in zip('ABCDEF',[10,10,42,14,14,22]):
        ws6.column_dimensions[w].width=width

    # Header
    for col,lbl in enumerate(['Akun','Waktu','Kombinasi Produk yang Dibeli',
                               'Jumlah Resi','Total Qty','Contoh No. Resi'],1):
        c=ws6.cell(row=1,column=col,value=lbl)
        c.fill=HDR2; c.font=HFNT2; c.alignment=C2; c.border=BD2
    ws6.row_dimensions[1].height=30

    resi_map2=dd2(list)
    for r in rows: resi_map2[r['no_resi']].append(r)

    # Warna
    F_SINGLE = PatternFill('solid',fgColor='E8F5E9')   # 1 produk
    F_MULTI  = PatternFill('solid',fgColor='FFEBEE')   # campur produk
    F_QTY    = PatternFill('solid',fgColor='FFF9C4')   # 1 produk qty>1

    row6=2
    akun_waktu_pairs=[]
    seen_aw=set()
    for r in rows:
        aw=(r.get('akun','-'),r.get('waktu','-'))
        if aw not in seen_aw: seen_aw.add(aw); akun_waktu_pairs.append(aw)
    # Jika tidak ada tag akun/waktu
    if not seen_aw or all(a=='-' for a,w in seen_aw):
        akun_waktu_pairs=[('-','-')]

    for akun,waktu in akun_waktu_pairs:
        if akun=='-':
            subset=resi_map2
        else:
            subset={k:v for k,v in resi_map2.items()
                    if v[0].get('akun','-')==akun and v[0].get('waktu','-')==waktu}
        if not subset: continue

        # Buat label kombinasi per resi
        combo_groups=dd2(list)  # label → [resi]
        for resi,items in subset.items():
            # Susun label: "SKU x qty" per item, diurutkan agar konsisten
            parts=[]
            for it in sorted(items, key=lambda x:x['sku']):
                nama=it['nama_produk'].rstrip(',').strip() if it.get('nama_produk') else it['sku']
                # Nama pendek: ambil 3 kata pertama
                short=' '.join(nama.split()[:4])
                parts.append(f"{short}  x{it['qty']}")
            label=' | '.join(parts)
            combo_groups[label].append(resi)

        # Urutkan dari jumlah resi terbanyak
        sorted_combos=sorted(combo_groups.items(), key=lambda x:-len(x[1]))

        for label,resi_list in sorted_combos:
            total_qty=sum(sum(r['qty'] for r in resi_map2[res]) for res in resi_list)
            contoh=resi_list[0]
            # Tentukan warna: campur = merah muda, qty>1 = kuning, single = hijau
            is_multi_prod='|' in label
            first_items=resi_map2[resi_list[0]]
            first_qty=sum(r['qty'] for r in first_items)
            if is_multi_prod: f=F_MULTI
            elif first_qty>1: f=F_QTY
            else: f=F_SINGLE

            vals=[akun,waktu,label,len(resi_list),total_qty,contoh]
            aligns=[C2,C2,L2,C2,C2,L2]
            for col,(v,al) in enumerate(zip(vals,aligns),1):
                c=ws6.cell(row=row6,column=col,value=v)
                c.fill=f; c.alignment=al; c.border=BD2
                c.font=Font(size=10,name='Calibri')
            # Jumlah resi dan qty bold
            ws6.cell(row=row6,column=4).font=Font(bold=True,size=12,name='Calibri')
            ws6.cell(row=row6,column=5).font=Font(bold=True,size=12,name='Calibri')
            # Campur produk: warnai merah pada kolom jumlah
            if is_multi_prod:
                ws6.cell(row=row6,column=4).font=Font(bold=True,size=12,color='CC0000',name='Calibri')
            row6+=1

        # Subtotal
        sub_resi=len(subset)
        sub_qty=sum(sum(r['qty'] for r in v) for v in subset.values())
        ws6.merge_cells(f'A{row6}:C{row6}')
        c=ws6.cell(row=row6,column=1,
                    value=f'  SUBTOTAL  {akun} {waktu}  —  {sub_resi} resi')
        c.fill=AMBER; c.font=Font(bold=True,size=10,name='Calibri')
        c.alignment=L2; c.border=BD2
        for col,val in [(4,sub_resi),(5,sub_qty)]:
            ct=ws6.cell(row=row6,column=col,value=val)
            ct.fill=AMBER; ct.font=Font(bold=True,size=12,name='Calibri')
            ct.alignment=C2; ct.border=BD2
        row6+=2

    wb.save(output_path)
    return output_path
